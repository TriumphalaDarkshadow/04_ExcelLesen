
from openpyxl import Workbook
from openpyxl import load_workbook
import os


#hole den pfad in dem sich die rechnungen befinden
directory = os.getcwd()+'/Rechnung/'

#hole liste von files im ordner Rechnung
folder = os.listdir(directory)


y = ' '
personen = []
briefumschlag_gesamt = 0
bleistift_gesamt = 0
lineal_gesamt = 0
textmarker_gesamt = 0


#laeuft durch alle Excel Dateien
for files in folder: 
    if files.endswith(".xlsx"):        
       
#        print files
        #wb = load_workbook(filename)
        #wb2 = load_workbook('Rechnung_EricIdle.xlsx')
        wb2 = load_workbook(os.path.join(directory, files))
        #print(wb2.sheetnames)
        sheets = wb2.worksheets
        
       
        #Hole namen aus dem ersten sheet eines workbooks
        vorname = sheets[0]['B3'].value
        nachname = sheets[0]['B4'].value
        name = (vorname, nachname)
        personen.append(name)

#       #laeuft durch Sheets
        for sheet in sheets:
        #    zaehler_dateien += 1
            
         
            #Hole anzahl der Artikel
            briefumschlag = sheet['C7'].value
            bleistift = sheet['C8'].value
            lineal = sheet['C9'].value
            textmarker = sheet['C10'].value
            
            #addiere zur gesamt anzahl
            briefumschlag_gesamt+= briefumschlag 
            bleistift_gesamt += bleistift 
            lineal_gesamt += lineal 
            textmarker_gesamt += textmarker 
            

                        
    


    else:
       # print("Datei konnte nicht gefunden werden")
       # print("\n")
        continue


#print(briefumschlag_gesamt)
#print(bleistift_gesamt)
#print(lineal_gesamt)
#print(textmarker_gesamt)


#        print("Erstelle neues workbook")
wb3 = Workbook()  
ws = wb3.active
ws.title ="Umsatz"

            #Kunden Namen einlesen
ws['A1'] = 'Es wurden '+ str(len(folder))  + ' Dateien eingelesen'
ws['A3'] = 'Artikel'
ws['A4'] = 'Briefumschlag'
ws['A5'] = 'Bleistift'
ws['A6'] = 'Lineael'
ws['A7'] = 'Textmarker'

ws['B3'] = 'Gesamtzahl'
ws['B4'] = briefumschlag_gesamt
ws['B5'] = bleistift_gesamt 
ws['B6'] = lineal_gesamt 
ws['B7'] = textmarker_gesamt 

wb3.save('Umsatz.xlsx')

print ("Es wurden", len(folder) , "Dateien eingelesen")
print ("\n")
print ("Artikel", "            ", "Gesamtzahl")
print ("Briefumschlag", "         ", briefumschlag_gesamt)
print ("Bleistift", "             ", bleistift_gesamt)
print ("Lineal", "                ", lineal_gesamt)
print ("Textmartker", "           ", textmarker_gesamt)
print ("\n")



kundenliste = input("Kundenliste ausgeben y/n ? ")
if kundenliste == 'y':
    def sort_by_lastname(e):
      return e[1]
    
    
    personen.sort(key=sort_by_lastname)
    for person in personen:
      print(person[0] + " "+ person[1])

else:
 exit()

